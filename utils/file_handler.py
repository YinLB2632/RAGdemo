import os
import hashlib
from utils.logger_handler import logger
from langchain_core.documents import Document
from langchain_community.document_loaders import (
    BSHTMLLoader,
    CSVLoader,
    Docx2txtLoader,
    PyPDFLoader,
    TextLoader,
)
from openpyxl import load_workbook


def get_file_md5_hex(filepath: str):     # 获取文件的md5的十六进制字符串

    if not os.path.exists(filepath):
        logger.error(f"[md5计算]文件{filepath}不存在")
        return

    if not os.path.isfile(filepath):
        logger.error(f"[md5计算]路径{filepath}不是文件")
        return

    md5_obj = hashlib.md5()

    chunk_size = 4096       # 4KB分片，避免文件过大爆内存
    try:
        with open(filepath, "rb") as f:     # 必须二进制读取
            while chunk := f.read(chunk_size):
                md5_obj.update(chunk)

            """
            chunk = f.read(chunk_size)
            while chunk:
                md5_obj.update(chunk)
                chunk = f.read(chunk_size)
            """
            md5_hex = md5_obj.hexdigest()
            return md5_hex
    except Exception as e:
        logger.error(f"计算文件{filepath}md5失败，{str(e)}")
        return None


def listdir_with_allowed_type(path: str, allowed_types: tuple[str]):        # 返回文件夹内的文件列表（允许的文件后缀）
    files = []

    if not os.path.isdir(path):
        logger.error(f"[listdir_with_allowed_type]{path}不是文件夹")
        return allowed_types

    normalized_types = tuple(f".{file_type.lstrip('.').lower()}" for file_type in allowed_types)

    for f in os.listdir(path):
        # Windows 上的大写后缀也需要按允许类型进行匹配。
        full_path = os.path.join(path, f)
        # 精确比较真实后缀可避免误匹配，并排除同名目录。
        if os.path.splitext(f)[1].lower() in normalized_types and os.path.isfile(full_path):
            files.append(full_path)

    return tuple(files)


def pdf_loader(filepath: str, passwd=None) -> list[Document]:
    return PyPDFLoader(filepath, passwd).load()


def txt_loader(filepath: str) -> list[Document]:
    return TextLoader(filepath, encoding="utf-8").load()


def markdown_loader(filepath: str) -> list[Document]:
    # Markdown 作为 UTF-8 文本文件加载。
    return TextLoader(filepath, encoding="utf-8").load()


def docx_loader(filepath: str) -> list[Document]:
    # 使用 docx2txt 解析 Word，单个文件失败由现有遍历逻辑继续处理。
    return Docx2txtLoader(filepath).load()


def csv_loader(filepath: str) -> list[Document]:
    # CSV 的每行会转换为可供后续分片的文档。
    return CSVLoader(filepath, encoding="utf-8").load()


def excel_loader(filepath: str) -> list[Document]:
    # Excel 仅需要读取单元格内容，使用 openpyxl 可避开 unstructured 及其 Office
    # 文件类型识别、OCR 等大型依赖链，降低部署依赖并避免无关组件缺失导致加载失败。
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    try:
        documents = []
        for worksheet in workbook.worksheets:
            rows = []
            for row in worksheet.iter_rows():
                values = [
                    str(cell.value).strip()
                    for cell in row
                    if cell.value is not None and str(cell.value).strip()
                ]
                if values:
                    rows.append(" | ".join(values))

            if rows:
                documents.append(Document(
                    page_content="\n".join(rows),
                    metadata={"source": filepath, "sheet_name": worksheet.title},
                ))

        return documents
    finally:
        # 只读工作簿仍会持有底层 ZIP 文件句柄，必须关闭以便 Windows 后续移动或删除文件。
        workbook.close()


def html_loader(filepath: str) -> list[Document]:
    # 中文政府和资讯网页通常为 UTF-8；显式指定编码并使用 html.parser，避免 Windows 默认 GBK 或 lxml 对不规范页面漏读正文。
    return BSHTMLLoader(
        filepath,
        open_encoding="utf-8",
        bs_kwargs={"features": "html.parser"},
        get_text_separator="\n",
    ).load()
